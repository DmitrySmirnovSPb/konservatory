from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Color, numbers,PatternFill, Border, Side, Alignment, GradientFill


class Excel(object):

    def __init__(self, link: str):
        self.link = link
        try:
            self.wb = load_workbook(self.link)
            self.error = None
        except AttributeError as e:
            print()
            exit(e)
# Инициация листа по названию sheetName
    def initSheet(self, sheetName: str):
        self.Sheet = self.getSheeltByName(sheetName)
        self.maxRow = self.Sheet.max_row
        self.maxColumn = self.Sheet.max_column
# Получить список листов в книге
    def getSheets(self):
        return self.wb.sheetnames
# Получить лист по названию
    def getSheeltByName(self, name: str):
        return self.wb[name]
# Получить контент с активного листа со строки startR по строку endR, колонки от startC по колонку endC
    def getContent(self, startR = 1, endR = 0, startC = 1, endC = 0, clear = False):
        rows = self.maxRow if endR == 0 or self.maxRow < endR else endR
        columns = self.maxColumn if endC == 0 or self.maxColumn < endC else endC
        result = {}
        for i in range(startR, rows + 1):
            result[i] = self.getLineContent(i, startC, columns, clear)
        return result
# Возвращает с листа Sheet строку с номером line со столбца с № start по № colums в формате словаря с ключами по № столбца
    def getLineContent(self, line, start, columns, clear):
        lineDict = {}
        for i in range(start, columns + 1):
            lineDict[i] = self.removeAllUnnecessarySpaces(self.getCell(line, i)) if clear else self.getCell(line, i)
        return lineDict
# Возвращает с листа Sheet колонку с номером line со строки с № start по № end в формате словаря с ключами по № столбца
    def getColumContent(self, column: int, start, end):
        columnDict = {}
        for i in range(start, end + 1):
            columnDict[i] = self.getCell(i, column)
        return columnDict
# возвращает значение ячейки № строки (r), № столбца (c) листа Sheet
    def getCell(self, r: int, c: int):
        return self.Sheet.cell(row=r, column=c).value
# возвращает значение ячейки в фомате А1 листа Sheet
    def getCellLitter(self, coordinates: str):
        return self.Sheet[coordinates].value
# Возвращает цвет шрифта в ячейке в формате FF000000
    def getFontColorCell(self, r: int, c: int):
        try:
            result = self.Sheet.cell(row=r, column=c).font.color.rgb
        except AttributeError as e:
            print(e)
            result = '000000'
        finally: return result

# Возврщает числовой формат ячейки с координатами r и c
    def getCellFormatNumber(self, r: int, c: int):
        return self.Sheet.cell(row=r, column=c).number_format

# Удалить лишние пробелы в string:str в начале, в конце и в середине
    def removeAllUnnecessarySpaces(self, string):
        if type(string) != str:
            return string
        while '  ' in string:
            string = string.replace('  ',' ')
        return string.strip()
# Добавить новую строку после последней сужествующей
    def addALine(self, data: dict):
        self.Sheet.append(data)
# Добавить новые строки в конец файла
    def addALines(self, data):
        for line in data:
            self.addALine(line)
# Записать в файл
    def saveFile(self, link = False):
        if link == False:
            link = self.link
        self.wb.save(link)